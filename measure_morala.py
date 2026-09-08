"""
measure_morala.py
─────────────────────────────────────────────────────────────────
将 measure.py 的实验替换为 Morala et al. (2021, Neural Networks)
"Towards a mathematical framework to inform neural network
modelling via polynomial regression" 风格的模拟实验，并保留
原项目中的几何距离测量（Mahalanobis / Procrustes distance）。

每个实验中同时构建三个模型并互相比较：
  (1) NN        : 单隐层（或多隐层）前馈神经网络（原始模型）
  (2) Taylor-PR : 按 Morala et al. 式(6)，从 NN 权重经泰勒展开
                  直接导出的多项式回归（仅单隐层时可用）
  (3) Full-PR   : 我原本的"全设定"多项式回归
                  （poly_interaction_features + 特殊特征 + 线性回归），
                  其最高阶 max_order = NN 隐层层数

输出内容：
  - 三个模型的 Test MSE（相对真实 y）
  - 模型两两之间预测的 MSE（论文的核心指标：PR 是否复现 NN）
  - 模型两两之间的 Mahalanobis 距离（预测向量）与
    Procrustes 形状距离（响应曲面构形 [X, ŷ]）
  - NN 训练过程中逐 epoch 的 线性层 vs 激活层 几何距离
  - 最后一个采集 epoch 的逐样本成对马氏距离

激活函数接口：--activation {softplus, tanh, sigmoid, relu, identity}
（relu 在 0 点不可导，Taylor-PR 会被自动跳过并提示）

用法示例：
  python measure_morala.py                          # 跑默认的全部实验组
  python measure_morala.py --case paper_poly2 \
      --activation softplus --h1 4 --q 3 --scaling -1,1
  python measure_morala.py --repeats 20             # 重复模拟看 MSE 分布
"""

import argparse
import math
from collections import defaultdict
from itertools import product, combinations_with_replacement

import numpy as np
import torch
# 多进程并行时，每个 worker 限制为单线程，避免 N 进程 × M 线程 过度抢占 CPU。
# （网络极小，单线程足够；并行收益来自进程级而非线程级。）
torch.set_num_threads(1)
import torch.nn as nn
import torch.optim as optim

from scipy.spatial import distance
from sklearn.covariance import LedoitWolf
from sklearn.decomposition import PCA
from sklearn.cross_decomposition import CCA
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import train_test_split

# 逐层泰勒展开模块：把任意（含多隐层）NN 展开成多项式，作为多隐层情形的对照模型
try:
    from taylor_expand import expand_nn_to_polynomial
    _HAS_TAYLOR_EXPAND = True
except ImportError:
    _HAS_TAYLOR_EXPAND = False

# 注：距离改用 scipy.spatial.procrustes（在 kendall_shape_distance 内 import），
# 不再依赖 geomstats，因此也不再需要 numpy<2 的版本约束。


# ─────────────────────────────────────────────
# 激活函数接口
# ─────────────────────────────────────────────
ACTIVATIONS = {
    "softplus": nn.Softplus,
    "tanh":     nn.Tanh,
    "sigmoid":  nn.Sigmoid,
    "relu":     nn.ReLU,     # 仅供 NN 使用；Taylor-PR 不支持（0 点不可导）
    "identity": nn.Identity,  # 线性/恒等激活，用作 NN 表达能力对照
    "identical": nn.Identity, # 常见误写别名，等价于 identity
}

ACTIVATION_HOOK_TYPES = (nn.ReLU, nn.Sigmoid, nn.Tanh, nn.Softplus, nn.Identity)


def activation_derivs_at_zero(act_name: str, q: int):
    """
    用自动微分精确计算 g^(n)(0)，n = 0..q。
    对应论文 2.2 节中各激活函数在 0 点的泰勒系数来源。
    """
    if act_name in {"identity", "identical"}:
        return [0.0, 1.0] + [0.0] * max(0, q - 1)
    if act_name == "relu":
        raise ValueError("ReLU 在 0 点不可导，无法做泰勒展开（参见论文结论：需分段多项式近似）。")
    act = ACTIVATIONS[act_name]()
    x = torch.zeros(1, dtype=torch.float64, requires_grad=True)
    g = act(x)
    derivs = [g.item()]
    cur = g
    for _ in range(q):
        (cur,) = torch.autograd.grad(cur.sum(), x, create_graph=True)
        derivs.append(cur.item())
    return derivs  # derivs[n] = g^(n)(0)


# ─────────────────────────────────────────────
# Hook 系统：按 epoch 采集线性层/激活层输出（沿用原项目）
# ─────────────────────────────────────────────
linear_by_epoch     = defaultdict(lambda: defaultdict(list))
activation_by_epoch = defaultdict(lambda: defaultdict(list))
current_epoch = -1
_capture_this_step = False


def reset_capture():
    global linear_by_epoch, activation_by_epoch, current_epoch, _capture_this_step
    linear_by_epoch     = defaultdict(lambda: defaultdict(list))
    activation_by_epoch = defaultdict(lambda: defaultdict(list))
    current_epoch = -1
    _capture_this_step = False


def to_cpu_detached(x):
    if torch.is_tensor(x):
        return x.detach().cpu()
    elif isinstance(x, (list, tuple)):
        return type(x)(to_cpu_detached(t) for t in x)
    elif isinstance(x, dict):
        return {k: to_cpu_detached(v) for k, v in x.items()}
    return x


def make_linear_hook(name):
    def hook(module, inputs, output):
        if _capture_this_step:
            linear_by_epoch[current_epoch][name].append(to_cpu_detached(output))
    return hook


def make_activation_hook(name):
    def hook(module, inputs, output):
        if _capture_this_step:
            activation_by_epoch[current_epoch][name].append(to_cpu_detached(output))
    return hook


def register_hooks(model):
    hooks = []
    for name, m in model.named_modules():
        if len(list(m.children())) == 0:
            if isinstance(m, nn.Linear):
                hooks.append(m.register_forward_hook(make_linear_hook(name)))
            elif isinstance(m, ACTIVATION_HOOK_TYPES):
                hooks.append(m.register_forward_hook(make_activation_hook(name)))
    return hooks


def remove_hooks(hooks):
    for h in hooks:
        h.remove()


# ─────────────────────────────────────────────
# 数据生成：Morala et al. 的模拟框架
# ─────────────────────────────────────────────
def poly_interaction_features(X, max_order=None):
    """（沿用原项目）生成所有总次数 1..max_order 的交互多项式特征"""
    n_samples, n_features = X.shape
    if max_order is None:
        max_order = 1
    features, powers = [], []
    for total_degree in range(1, max_order + 1):
        for exponents in product(range(total_degree + 1), repeat=n_features):
            if sum(exponents) == total_degree:
                features.append(np.prod(X ** exponents, axis=1))
                powers.append(exponents)
    return np.column_stack(features), powers


def make_paper_polynomial(n=200, p=3, degree=2, noise_sd=0.1, rng=0):
    """
    论文 3 节通用框架：
      Xi ~ N(μi, 1)，μi ~ U(-10, 10)；
      Y = 完整 degree 阶多项式（含交互项），β ~ U(-5, 5)；
      再加 N(0, 0.1) 噪声。
    返回未缩放的 (X, y, true_betas)。
    """
    g = np.random.default_rng(rng)
    mus = g.uniform(-10, 10, size=p)
    X = g.normal(loc=mus, scale=1.0, size=(n, p))
    design, powers = poly_interaction_features(X, max_order=degree)
    betas = g.uniform(-5, 5, size=design.shape[1])
    beta0 = g.uniform(-5, 5)
    y = beta0 + design @ betas + g.normal(0, noise_sd, size=n)
    return X, y, (beta0, betas, powers)


def make_smooth_nonlinear_lowdim(n=200, p=3, noise_sd=0.1, rng=0, randomize=False):
    """
    非多项式真值。
      固定版 (randomize=False)：y = sin(x1) + x2² + x2·x3 + x3 + ε
      随机版 (randomize=True) ：每个 seed 重新抽 sin 的频率/相位/幅度，
          y = a·sin(ω·x1 + φ) + x2² + x2·x3 + x3 + ε
      随机版的意义（建议一）：特征族里的 sin(x_i) 频率固定为 1、无相位，
          与数据中随机的 ω、φ 对不上，因此 Full-PR 无论如何都无法靠
          sin 特殊项"作弊"命中——这是无法被特征族补上的严格欠设定。
    """
    g = np.random.default_rng(rng)
    X = g.normal(0, 1, size=(n, p))
    if randomize:
        omega = g.uniform(1.5, 3.5)
        phi   = g.uniform(0, 2 * np.pi)
        amp   = g.uniform(1.0, 2.5)
        s = amp * np.sin(omega * X[:, 0] + phi)
    else:
        s = np.sin(X[:, 0])
    y = s + X[:, 1] ** 2 + X[:, 1] * X[:, 2] + X[:, 2] + g.normal(0, noise_sd, size=n)
    return X, y, None


def scale_minmax(train, test, interval=(-1.0, 1.0)):
    """按训练集逐列 min-max 缩放到给定区间（论文中对 X 和 y 都缩放）"""
    lo, hi = interval
    mn = train.min(axis=0)
    mx = train.max(axis=0)
    span = np.where(mx - mn == 0, 1.0, mx - mn)

    def _s(z):
        return lo + (hi - lo) * (z - mn) / span

    return _s(train), _s(test)


def generate_dataset(case, n=200, p=3, rng=0, scaling=(-1.0, 1.0),
                     scale_inputs=True):
    if case == "paper_poly2":
        X, y, _ = make_paper_polynomial(n=n, p=p, degree=2, rng=rng)
    elif case == "paper_poly3":
        X, y, _ = make_paper_polynomial(n=n, p=p, degree=3, rng=rng)
    elif case == "paper_poly4":
        # 4 阶 > 最大隐层数(3) → 所有深度的 Full-PR 都欠设定，无一命中
        X, y, _ = make_paper_polynomial(n=n, p=p, degree=4, rng=rng)
    elif case == "smooth_nonlinear":
        X, y, _ = make_smooth_nonlinear_lowdim(n=n, p=p, rng=rng)
    elif case == "smooth_nonlinear_rand":
        # 频率/相位随机：sin 特殊项无法命中，严格欠设定（建议一）
        X, y, _ = make_smooth_nonlinear_lowdim(n=n, p=p, rng=rng, randomize=True)
    else:
        raise ValueError(f"未知数据集：{case}")

    # 论文流程：先缩放（X 与 y 都缩放），再 75/25 划分。
    # scale_inputs=False 时不缩放 X（y 仍缩放以稳定训练）——
    # 用于"未缩放输入"那一遍：激活层在更宽的工作范围上，非线性更易显现。
    X_tr, X_te, y_tr, y_te = train_test_split(
        X, y.reshape(-1, 1), test_size=0.25, random_state=42
    )
    if scale_inputs:
        X_tr, X_te = scale_minmax(X_tr, X_te, scaling)
    y_tr, y_te = scale_minmax(y_tr, y_te, scaling)

    to_t = lambda a: torch.tensor(a, dtype=torch.float32)
    return to_t(X_tr), to_t(X_te), to_t(y_tr), to_t(y_te)


# ─────────────────────────────────────────────
# 神经网络（可配置隐层与激活函数）
# ─────────────────────────────────────────────
class ConfigurableNet(nn.Module):
    """
    隐层宽度由 hidden_layers 列表给出，激活函数由 act_name 给出，
    输出层为线性（满足论文回归设定）。
    nn.Sequential 命名为 model.0, model.1, ... 与原 hook 系统兼容。
    """
    def __init__(self, input_size, hidden_layers, act_name):
        super().__init__()
        act_cls = ACTIVATIONS[act_name]
        layers, prev = [], input_size
        for h in hidden_layers:
            layers.append(nn.Linear(prev, h))
            layers.append(act_cls())
            prev = h
        layers.append(nn.Linear(prev, 1))
        self.model = nn.Sequential(*layers)

    def forward(self, x):
        return self.model(x)


def train_nn(model, X_train, y_train, X_eval=None, y_eval=None, epochs=1500,
             optimizer_name="rprop", lr=0.01, capture_every=150, verbose=True):
    """
    全批量训练（模仿论文用 neuralnet 的 Rprop 弹性反向传播）。
    每隔 capture_every 个 epoch：
      - 通过 hook 采集一次各层输出（供几何距离分析）
      - 同步记录 Train MSE（及给定 X_eval 时的 Test MSE）
    返回 history: {epoch: {"train_mse":…, "test_mse":…}}，
    其 epoch 与几何距离的采集 epoch 一一对应，方便对比。
    """
    global current_epoch, _capture_this_step
    if optimizer_name == "rprop":
        opt = optim.Rprop(model.parameters(), lr=lr)
    elif optimizer_name == "adam":
        opt = optim.Adam(model.parameters(), lr=1e-3)
    else:
        raise ValueError("optimizer 仅支持 rprop / adam")

    criterion = nn.MSELoss()
    history = {}
    model.train()
    for epoch in range(epochs):
        current_epoch = epoch
        is_capture = (epoch % capture_every == 0) or (epoch == epochs - 1)
        _capture_this_step = is_capture
        out = model(X_train)
        loss = criterion(out, y_train)
        opt.zero_grad()
        loss.backward()
        opt.step()
        _capture_this_step = False     # 评估 Test MSE 时不再触发 hook 采集

        if is_capture:
            rec = {"train_mse": loss.item()}
            if X_eval is not None:
                model.eval()
                with torch.no_grad():
                    rec["test_mse"] = criterion(model(X_eval), y_eval).item()
                model.train()
            history[epoch] = rec

        if verbose and (epoch + 1) % max(1, epochs // 5) == 0:
            print(f"    Epoch {epoch + 1}/{epochs}, Train MSE: {loss.item():.6f}")
    return history


# ─────────────────────────────────────────────
# Morala et al. 式(6)：NN 权重 → 多项式系数
# ─────────────────────────────────────────────
def all_multi_indices(p, q):
    """枚举所有 1 ≤ Σmi ≤ q 的多重指数 (m1,...,mp)"""
    for t in range(1, q + 1):
        for combo in combinations_with_replacement(range(p), t):
            m = [0] * p
            for idx in combo:
                m[idx] += 1
            yield tuple(m)


def nn_to_taylor_polynomial(model, act_name, q):
    """
    实现论文式 (5)(6)：
      β0          = v0 + Σ_j v_j Σ_{n=0}^{q} g^(n)(0)/n! · w0j^n
      β_{l1...lt} = Σ_j v_j Σ_{n=t}^{q} g^(n)(0) /
                    ((n-t)!·m1!···mp!) · w0j^{n-t} · w1j^{m1}···wpj^{mp}
    其中 w0j 为隐层偏置，v0 为输出层偏置。仅适用于单隐层网络。
    返回 (beta0, {multi_index: beta})。
    """
    linears = [m for m in model.modules() if isinstance(m, nn.Linear)]
    if len(linears) != 2:
        raise ValueError("Taylor-PR 仅支持单隐层网络（论文设定）。")

    W  = linears[0].weight.detach().double().numpy()          # (h1, p)
    b  = linears[0].bias.detach().double().numpy()            # (h1,)  = w0j
    v  = linears[1].weight.detach().double().numpy().ravel()  # (h1,)
    v0 = float(linears[1].bias.detach().double().numpy()[0])

    g = activation_derivs_at_zero(act_name, q)                # g[n] = g^(n)(0)
    h1, p = W.shape

    # 截距
    beta0 = v0 + sum(
        v[j] * sum(g[n] / math.factorial(n) * b[j] ** n for n in range(q + 1))
        for j in range(h1)
    )

    # 各阶系数
    betas = {}
    for m in all_multi_indices(p, q):
        t = sum(m)
        denom_m = np.prod([math.factorial(mi) for mi in m])
        coef = 0.0
        for j in range(h1):
            wm = np.prod([W[j, i] ** m[i] for i in range(p) if m[i] > 0])
            s = sum(g[n] / math.factorial(n - t) * b[j] ** (n - t)
                    for n in range(t, q + 1))
            coef += v[j] * wm * s
        betas[m] = coef / denom_m
    return beta0, betas


def eval_taylor_polynomial(beta0, betas, X):
    X = np.asarray(X, dtype=np.float64)
    pred = np.full(X.shape[0], beta0)
    for m, beta in betas.items():
        term = np.ones(X.shape[0])
        for i, mi in enumerate(m):
            if mi > 0:
                term = term * X[:, i] ** mi
        pred += beta * term
    return pred.reshape(-1, 1)


def _taylor_sanity_check(model, act_name, q, beta0, betas, X, tol=1e-6):
    """
    验证：多项式预测 == 直接用截断泰勒级数算激活的 NN 前向。
    二者在数值精度内应一致（这是式(6)推导的恒等式）。
    """
    linears = [m for m in model.modules() if isinstance(m, nn.Linear)]
    W  = linears[0].weight.detach().double().numpy()
    b  = linears[0].bias.detach().double().numpy()
    v  = linears[1].weight.detach().double().numpy().ravel()
    v0 = float(linears[1].bias.detach().double().numpy()[0])
    g  = activation_derivs_at_zero(act_name, q)

    X = np.asarray(X, dtype=np.float64)
    U = X @ W.T + b                                   # 突触电位 u_j
    Yh = sum(g[n] / math.factorial(n) * U ** n for n in range(q + 1))
    z_direct = v0 + Yh @ v
    z_poly = eval_taylor_polynomial(beta0, betas, X).ravel()
    err = np.max(np.abs(z_direct - z_poly))
    assert err < tol, f"Taylor-PR 自检失败，max|diff|={err:.3e}"
    return err


def synaptic_potential_coverage(model, X, act_name, q, err_threshold=0.1):
    """
    论文图 5(C) 的数值版诊断：统计突触电位 u_j 落在
    泰勒近似误差 < err_threshold 区间内的比例。
    """
    linears = [m for m in model.modules() if isinstance(m, nn.Linear)]
    W = linears[0].weight.detach().double().numpy()
    b = linears[0].bias.detach().double().numpy()
    U = np.asarray(X, dtype=np.float64) @ W.T + b     # (n, h1)

    act = ACTIVATIONS[act_name]()
    g = activation_derivs_at_zero(act_name, q)
    grid = np.linspace(-6, 6, 2001)
    with torch.no_grad():
        true_vals = act(torch.tensor(grid)).numpy()
    taylor_vals = sum(g[n] / math.factorial(n) * grid ** n for n in range(q + 1))
    ok = np.abs(true_vals - taylor_vals) < err_threshold
    # 找包含 0 的最大连续可接受区间
    zero_idx = np.searchsorted(grid, 0.0)
    left = zero_idx
    while left > 0 and ok[left - 1]:
        left -= 1
    right = zero_idx
    while right < len(grid) - 1 and ok[right + 1]:
        right += 1
    lo, hi = grid[left], grid[right]
    inside = float(((U >= lo) & (U <= hi)).mean())
    return inside, (lo, hi)


# ─────────────────────────────────────────────
# 我原本的"全设定"多项式模型（Full-PR）
# ─────────────────────────────────────────────
def custom_features_full(X, max_order=None, include_special=True, full_poly=False):
    """
    全设定多项式特征（按用户规格）：
      - 所有总次数 1..max_order 的完整交互多项式项
      - 特殊项：每个变量的 sin(x_i) 与 e^{x_i}
    max_order 由调用方传入（实验中 = NN 隐层层数）。
    """
    base_features, _ = poly_interaction_features(X, max_order)
    all_features = [base_features]

    if include_special:
        special = []
        for i in range(X.shape[1]):
            xi = X[:, i]
            special.extend([
                np.sin(xi),
                np.exp(np.clip(xi, -30, 30)),
            ])
        all_features.append(np.column_stack(special))

    combined = np.column_stack(all_features)
    if full_poly:
        expanded, _ = poly_interaction_features(combined, max_order)
        return expanded
    return combined


def fit_full_pr(X_train, y_train, X_test, max_order, include_special=True):
    F_tr = custom_features_full(X_train, max_order=max_order,
                                include_special=include_special)
    F_te = custom_features_full(X_test, max_order=max_order,
                                include_special=include_special)
    lr = LinearRegression().fit(F_tr, y_train)
    return lr.predict(F_te), lr


# ─────────────────────────────────────────────
# 几何距离函数（沿用原项目）
# ─────────────────────────────────────────────
def pooled_covariance(A, B, shrinkage=None, ridge=0.0):
    X = np.vstack([A, B])
    if shrinkage is None:
        cov = np.cov(X, rowvar=False)
    elif shrinkage == 'ledoitwolf':
        cov = LedoitWolf().fit(X).covariance_
    else:
        raise ValueError("Unsupported shrinkage")
    cov = np.atleast_2d(cov)
    if ridge > 0:
        cov = cov + ridge * np.eye(cov.shape[0])
    VI = np.linalg.pinv(cov)
    return cov, VI


def mahalanobis(A, B, shrinkage='ledoitwolf', ridge=0.0):
    A = np.atleast_2d(A); B = np.atleast_2d(B)
    muA, muB = A.mean(axis=0), B.mean(axis=0)
    _, VI = pooled_covariance(A, B, shrinkage=shrinkage, ridge=ridge)
    return distance.mahalanobis(muA, muB, VI)


def kendall_shape_distance(A, B, n_landmarks=16, atol=1e-12):
    """
    Procrustes 形状距离（替换原 Kendall 测地距离）。

    用 scipy.spatial.procrustes：对两个配置做最优平移/缩放/旋转对齐后，
    返回 disparity = 对齐后逐点残差的平方和 M² = Σ‖a_i − b_i‖²。

    与原 Kendall 版的两点区别：
      1. 使用全部 n 个点参与对齐，而非只取 PCA 投影的前 n_comp 行做地标
         —— 因此尾部/极端样本不再被排除在计算之外；
      2. procrustes 要求两个矩阵形状相同。当 A、B 列数不同（如激活层 vs
         下一线性层），先各自 PCA 到相同的 n_comp 维再对齐。
    （函数名/签名保持不变，所有调用点无需改动。）
    """
    from scipy.spatial import procrustes

    A = np.asarray(A, dtype=np.float64)
    B = np.asarray(B, dtype=np.float64)
    if A.shape[0] < 2 or A.shape[1] < 1 or B.shape[1] < 1:
        return float("nan")

    # 列数不同时：各自 PCA 降到相同维度；列数相同则直接用原始坐标
    if A.shape[1] != B.shape[1]:
        n_comp = min(A.shape[1], B.shape[1], A.shape[0] - 1)
        if n_comp < 1:
            return float("nan")
        A = PCA(n_components=n_comp).fit_transform(A)
        B = PCA(n_components=n_comp).fit_transform(B)

    # 一维列无法用 procrustes 标准化（norm 后退化），补一列零升到二维
    if A.shape[1] == 1:
        A = np.hstack([A, np.zeros((A.shape[0], 1))])
        B = np.hstack([B, np.zeros((B.shape[0], 1))])

    # procrustes 会对每个输入做中心化 + Frobenius 归一化；
    # 若某个配置本身是常数（norm=0）会报错，先挡掉
    if np.allclose(A - A.mean(0), 0, atol=atol) or \
       np.allclose(B - B.mean(0), 0, atol=atol):
        return float("nan")

    _, _, disparity = procrustes(A, B)
    return max(float(disparity), 0.0)


def surface_shape_distance(X, predA, predB, n_landmarks=16):
    """
    模型间的响应曲面形状距离：
    把两模型的拟合图象 {(x_i, ŷ_i)} 视为 (n, p+1) 构形，
    再用 Procrustes 形状距离比较。
    """
    GA = np.hstack([X, predA.reshape(-1, 1)])
    GB = np.hstack([X, predB.reshape(-1, 1)])
    return kendall_shape_distance(GA, GB, n_landmarks=n_landmarks)


def activation_io_distances(model, X):
    """
    在【已训练好】的网络上，对每一个激活层，测量
        该层激活前的线性输出 u = (上一层输出·W + b)
        与激活后 g(u)
    之间的 Procrustes 形状距离。每一激活层单独给一个值。

    直接前向一次、用 hook 抓每个激活模块的 input/output，
    因此 u 与 g(u) 必然同形状(逐元素激活)，可直接比较，无需 PCA。
    返回 {层名: 距离}，键按激活层在网络中的顺序排列。
    """
    captured = []  # [(activation_module_name, u, g(u)), ...]
    handles = []

    def make_hook(name):
        def hook(module, inp, out):
            u  = inp[0].detach().cpu().numpy()   # 激活前（线性层输出）
            gu = out.detach().cpu().numpy()       # 激活后
            captured.append((name, u, gu))
        return hook

    act_types = (nn.ReLU, nn.Sigmoid, nn.Tanh, nn.Softplus)
    for name, m in model.named_modules():
        if isinstance(m, act_types):
            handles.append(m.register_forward_hook(make_hook(name)))

    model.eval()
    with torch.no_grad():
        model(X if torch.is_tensor(X) else torch.tensor(X, dtype=torch.float32))
    for h in handles:
        h.remove()

    captured.sort(key=lambda t: int(t[0].split(".")[1]))
    return {name: kendall_shape_distance(u, gu) for name, u, gu in captured}


def next_name(name):
    k = int(name.split(".")[1])
    return f"model.{k + 1}"


def _pooled_cov_inv(X, Y, shrinkage='ledoitwolf', ridge=0.0):
    Z = np.vstack([X, Y])
    if shrinkage is None:
        cov = np.cov(Z, rowvar=False)
    elif shrinkage == 'ledoitwolf':
        cov = LedoitWolf().fit(Z).covariance_
    else:
        raise ValueError("shrinkage must be None or 'ledoitwolf'")
    cov = np.atleast_2d(cov)
    if ridge > 0:
        cov = cov + ridge * np.eye(cov.shape[0])
    return np.linalg.pinv(cov)


def _paired_mahalanobis(A, B, VI):
    D = A - B
    DV = D @ VI
    d2 = np.maximum(np.sum(DV * D, axis=1), 0.0)
    return np.sqrt(d2)


def paired_md_from_first(linear_concat, activation_concat,
                         source_name="model.0", align="auto",
                         k=None, shrinkage="ledoitwolf", ridge=0.0):
    assert source_name in linear_concat, f"{source_name} 不在 linear_concat 中"
    src = linear_concat[source_name]
    N, d_src = src.shape

    all_layers = {**linear_concat, **activation_concat}
    items = [(n, X) for n, X in all_layers.items() if n != source_name]
    items.sort(key=lambda pair: int(pair[0].split(".")[1]))

    results, summary = {}, {}
    for name, tgt in items:
        if tgt.shape[0] != N:
            raise ValueError(f"{name} 样本数不一致：{tgt.shape[0]} vs {N}")
        d_tgt = tgt.shape[1]
        mode = align
        if mode == "auto":
            mode = "match" if d_src == d_tgt else "cca"
        if mode == "match":
            if d_src != d_tgt:
                continue
            A_aligned, B_aligned = src, tgt
        else:
            k_eff = k or min(d_src, d_tgt, max(N - 1, 1))
            cca = CCA(n_components=k_eff, max_iter=5000)
            A_aligned, B_aligned = cca.fit_transform(src, tgt)

        VI = _pooled_cov_inv(A_aligned, B_aligned, shrinkage=shrinkage, ridge=ridge)
        dvec = _paired_mahalanobis(A_aligned, B_aligned, VI)
        results[name] = dvec
        summary[name] = (float(dvec.mean()), float(dvec.std()))
    return results, summary


def _fmt(x, w=12, prec=6):
    """统一的数值格式化：nan → '—'，否则定宽小数"""
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "—".center(w)
    return f"{x:>{w}.{prec}f}"


def epoch_layer_distances(X_input=None, verbose=True):
    """
    新指标：NN 中每一层输出与其上一层输出之间的 Procrustes 形状距离，
    并对每个 epoch 求所有相邻层对的均值（adjacent-layer shape mean）。

    - X_input 不为 None 时，把网络输入作为第 0 层，参与 input → model.0 的比较；
    - 同维的相邻层对（线性层 → 其激活层）同时给出 Mahalanobis 距离；
      维度不同的层对（激活层 → 下一线性层）Mahalanobis 不可比，记为 —；
    - 一维输出层无法构成形状空间，shape 记为 —，不计入均值。
    """
    linear_epoch_concat, activation_epoch_concat = {}, {}
    for ep, layer_dict in linear_by_epoch.items():
        linear_epoch_concat[ep] = {
            n: torch.cat(outs, dim=0).numpy() for n, outs in layer_dict.items()
        }
    for ep, layer_dict in activation_by_epoch.items():
        activation_epoch_concat[ep] = {
            n: torch.cat(outs, dim=0).numpy() for n, outs in layer_dict.items()
        }

    epoch_rows, epoch_shape_mean, epoch_io_shape = {}, {}, {}
    epoch_act_shape_mean = {}
    for ep in sorted(linear_epoch_concat.keys()):
        all_layers = {**linear_epoch_concat[ep],
                      **activation_epoch_concat.get(ep, {})}
        ordered = sorted(all_layers.items(),
                         key=lambda kv: int(kv[0].split(".")[1]))
        seq = ([("input", np.asarray(X_input))] if X_input is not None else []) \
              + ordered

        rows, shape_vals, act_shape_vals = [], [], []
        act_names = set(activation_epoch_concat.get(ep, {}).keys())
        for (n_prev, A), (n_cur, B) in zip(seq[:-1], seq[1:]):
            sd = kendall_shape_distance(A, B)
            md = (mahalanobis(A, B, shrinkage='ledoitwolf', ridge=1e-6)
                  if A.shape[1] == B.shape[1] else float("nan"))
            rows.append((f"{n_prev} → {n_cur}", md, sd))
            if not math.isnan(sd):
                shape_vals.append(sd)
                # 仅"线性层 → 其激活层"的对：逐元素激活在线性区内
                # 近似为相似变换（统一缩放+平移），shape 距离 → 0；
                # 因此这个子均值是对"该层注入了多少非线性"的直接计量。
                if n_cur in act_names:
                    act_shape_vals.append(sd)

        epoch_rows[ep] = rows
        epoch_shape_mean[ep] = (float(np.mean(shape_vals))
                                if shape_vals else float("nan"))
        epoch_act_shape_mean[ep] = (float(np.mean(act_shape_vals))
                                    if act_shape_vals else float("nan"))

        # 输入 ↔ 输出 的 shape 距离：
        # 网络输出是一维标量，本身构不成形状空间，因此沿用模型间比较的做法，
        # 用响应曲面构形 [X | z] 代表"输出端"几何，与输入构形 X 比较。
        # 它衡量整张网络对输入几何的总扭曲量（相邻层距离衡量的是逐层扭曲）。
        if X_input is not None and ordered:
            z_out = ordered[-1][1]            # 最后一层（线性输出，n×1）
            Xin = np.asarray(X_input)
            epoch_io_shape[ep] = kendall_shape_distance(
                Xin, np.hstack([Xin, z_out]))
        else:
            epoch_io_shape[ep] = float("nan")

    if verbose:
        print(f"\n  ┌{'─' * 66}")
        print(f"  │ NN 内部：逐 Epoch 相邻层几何距离")
        print(f"  ├{'─' * 66}")
        print(f"  │ {'Epoch':>6}  {'相邻层对':<22}{'Mahalanobis':>13}"
              f"{'Procrustes':>15}")
        for ep in sorted(epoch_rows.keys()):
            for i, (pair, md, sd) in enumerate(epoch_rows[ep]):
                ep_str = f"{ep:>6}" if i == 0 else " " * 6
                print(f"  │ {ep_str}  {pair:<22}{_fmt(md, 13)}{_fmt(sd, 15)}")
            print(f"  │ {' ' * 6}  {'★ 相邻层 shape 均值':<22}{' ' * 13}"
                  f"{_fmt(epoch_shape_mean[ep], 15)}")
            print(f"  │ {' ' * 6}  {'★ 激活层对均值(非线性)':<22}{' ' * 13}"
                  f"{_fmt(epoch_act_shape_mean[ep], 15)}")
            print(f"  │ {' ' * 6}  {'★ 输入 ↔ 输出 shape':<22}{' ' * 13}"
                  f"{_fmt(epoch_io_shape[ep], 15)}")
        print(f"  └{'─' * 66}")

    return (epoch_rows, epoch_shape_mean, epoch_act_shape_mean, epoch_io_shape,
            linear_epoch_concat, activation_epoch_concat)


# ─────────────────────────────────────────────
# 单次实验
# ─────────────────────────────────────────────
def run_experiment(case="paper_poly2", activation="softplus",
                   hidden_layers=(4,), q=3, scaling=(-1.0, 1.0),
                   n=200, p=3, epochs=1500, optimizer="rprop",
                   capture_every=150, seed=0, data_seed=None, init_seed=None,
                   include_special=True, verbose=True):
    # 拆分随机性来源：
    #   data_seed 控制数据生成（哪份数据集）
    #   init_seed 控制网络权重初始化（训练随机性）
    # 不指定时退回旧行为（都用 seed），保持向后兼容。
    if data_seed is None:
        data_seed = seed
    if init_seed is None:
        init_seed = seed

    torch.manual_seed(init_seed)
    np.random.seed(init_seed)
    reset_capture()

    hidden_layers = list(hidden_layers)
    n_hidden = len(hidden_layers)
    max_order_full_pr = n_hidden        # 要求：Full-PR 最高阶 = NN 隐层层数

    if verbose:
        print(f"\n{'=' * 72}")
        sp = "含sin/e^x" if include_special else "纯多项式"
        print(f"实验：case={case}  activation={activation}  "
              f"hidden={hidden_layers}  q={q}  scaling={list(scaling)}  "
              f"data_seed={data_seed} init_seed={init_seed}  Full-PR={sp}")
        print(f"{'=' * 72}")

    # 1. 数据（用 data_seed）
    X_tr, X_te, y_tr, y_te = generate_dataset(case, n=n, p=p, rng=data_seed,
                                              scaling=scaling)
    X_te_np, y_te_np = X_te.numpy(), y_te.numpy()

    # 2. 训练 NN（权重初始化用 init_seed；带 hook 采集）
    model = ConfigurableNet(p, hidden_layers, activation)
    hooks = register_hooks(model)
    history = train_nn(model, X_tr, y_tr, X_eval=X_te, y_eval=y_te,
                       epochs=epochs, optimizer_name=optimizer,
                       capture_every=capture_every, verbose=verbose)
    model.eval()
    with torch.no_grad():
        pred_nn = model(X_te).numpy()
    nn_mse = mean_squared_error(y_te_np, pred_nn)
    remove_hooks(hooks)

    res = {"case": case, "activation": activation, "h": tuple(hidden_layers),
           "q": q, "seed": seed, "data_seed": data_seed, "init_seed": init_seed,
           "nn_mse": nn_mse}

    # 2b. 每一激活层 u → g(u) 的 Procrustes 距离（在已训练好的网络上，缩放输入）
    #     测量"数据经过每个激活层时被改变了多少几何"。
    act_io_scaled = activation_io_distances(model, X_te)
    for name, d in act_io_scaled.items():
        res[f"act_io_scaled[{name}]"] = d
    res["act_io_scaled_mean"] = (
        float(np.nanmean(list(act_io_scaled.values()))) if act_io_scaled
        else float("nan"))
    res["act_io_scaled_sum"] = (
        float(np.nansum(list(act_io_scaled.values()))) if act_io_scaled
        else float("nan"))

    # 2c. 未缩放输入：在未缩放 X 上重新训练一个同结构网络，再测同样的逐层距离。
    #     （中间层无法"只换输入不换权重"，故必须重训以保持自洽。）
    torch.manual_seed(init_seed)
    np.random.seed(init_seed)
    Xu_tr, Xu_te, yu_tr, yu_te = generate_dataset(
        case, n=n, p=p, rng=data_seed, scaling=scaling, scale_inputs=False)
    model_u = ConfigurableNet(p, hidden_layers, activation)
    train_nn(model_u, Xu_tr, yu_tr, epochs=epochs,
             optimizer_name=optimizer, capture_every=epochs + 1, verbose=False)
    act_io_unscaled = activation_io_distances(model_u, Xu_te)
    for name, d in act_io_unscaled.items():
        res[f"act_io_unscaled[{name}]"] = d
    res["act_io_unscaled_mean"] = (
        float(np.nanmean(list(act_io_unscaled.values()))) if act_io_unscaled
        else float("nan"))
    res["act_io_unscaled_sum"] = (
        float(np.nansum(list(act_io_unscaled.values()))) if act_io_unscaled
        else float("nan"))

    if verbose:
        print(f"\n  ┌{'─' * 66}")
        print(f"  │ 每层激活 u → g(u) 的 Procrustes 距离（已训练网络）")
        print(f"  ├{'─' * 66}")
        print(f"  │ {'激活层':<14}{'缩放输入':>14}{'未缩放输入':>14}")
        names = sorted(set(act_io_scaled) | set(act_io_unscaled),
                       key=lambda s: int(s.split('.')[1]))
        for nm in names:
            print(f"  │ {nm:<14}{_fmt(act_io_scaled.get(nm, float('nan')), 14)}"
                  f"{_fmt(act_io_unscaled.get(nm, float('nan')), 14)}")
        print(f"  │ {'★ 均值':<14}{_fmt(res['act_io_scaled_mean'], 14)}"
              f"{_fmt(res['act_io_unscaled_mean'], 14)}")
        print(f"  └{'─' * 66}")

    # 3. Taylor-PR（Morala et al. 式(6)，仅单隐层 + 光滑激活）
    pred_tpr = None
    if n_hidden == 1 and activation != "relu":
        beta0, betas = nn_to_taylor_polynomial(model, activation, q)
        chk = _taylor_sanity_check(model, activation, q, beta0, betas, X_te_np)
        pred_tpr = eval_taylor_polynomial(beta0, betas, X_te_np)
        res["taylor_check_err"] = chk
        res["tpr_mse_vs_y"]  = mean_squared_error(y_te_np, pred_tpr)
        res["tpr_mse_vs_nn"] = mean_squared_error(pred_nn, pred_tpr)  # 论文核心指标
        cov_ratio, (lo, hi) = synaptic_potential_coverage(
            model, X_te_np, activation, q)
        res["u_coverage"] = cov_ratio
        if verbose:
            print(f"\n  [Taylor-PR] 多项式项数: {len(betas)}（恒等自检 "
                  f"max|diff|={chk:.2e}）")
            print(f"  [Taylor-PR] 突触电位落在可接受近似区间 "
                  f"[{lo:.2f}, {hi:.2f}] 内的比例: {cov_ratio:.1%}")
    else:
        reason = ("ReLU 在 0 点不可导" if activation == "relu"
                  else f"网络有 {n_hidden} 个隐层（式(6)仅适用单隐层）")
        if verbose:
            print(f"\n  [Taylor-PR] 跳过：{reason}")

    # 3b. LayerTaylor-PR（逐层泰勒展开，适用任意层数，含多隐层）
    #     当单隐层 Taylor-PR 不可用（多隐层）时，用 taylor_expand 模块把整网
    #     逐层展开成多项式，作为多隐层情形下的多项式对照模型。
    #     在数据点（突触电位均值）处展开，精度 order=q，总次数封顶防爆。
    pred_ltpr = None
    if _HAS_TAYLOR_EXPAND and activation != "relu":
        try:
            ltpr_poly = expand_nn_to_polynomial(
                model, input_dim=p, order=q, expansion_point="data",
                X_ref=X_tr.numpy(), max_total_degree=max(q, n_hidden * 2))
            pred_ltpr = ltpr_poly.predict(X_te_np)
            res["ltpr_mse_vs_y"]  = mean_squared_error(y_te_np, pred_ltpr)
            res["ltpr_mse_vs_nn"] = mean_squared_error(pred_nn, pred_ltpr)
            res["ltpr_n_terms"]   = ltpr_poly.n_terms
            res["ltpr_max_degree"] = ltpr_poly.max_total_degree
            if verbose:
                print(f"  [LayerTaylor-PR] 逐层展开：{ltpr_poly.n_terms} 项，"
                      f"最高 {ltpr_poly.max_total_degree} 次，"
                      f"MSE vs NN={res['ltpr_mse_vs_nn']:.3e}")
        except Exception as e:
            if verbose:
                print(f"  [LayerTaylor-PR] 展开失败：{e}")

    # 4. Full-PR（多项式模型，max_order = 隐层数）
    #    include_special=True：完整交互项 + sin(x_i) + e^{x_i}
    #    include_special=False：纯多项式（建议二，对 4 阶数据是数学上无争议的欠设定）
    res["fpr_special"] = include_special
    pred_fpr, fpr_model = fit_full_pr(X_tr.numpy(), y_tr.numpy(), X_te_np,
                                      max_order=max_order_full_pr,
                                      include_special=include_special)
    res["fpr_mse_vs_y"]  = mean_squared_error(y_te_np, pred_fpr)
    res["fpr_mse_vs_nn"] = mean_squared_error(pred_nn, pred_fpr)

    # 4b. 多项式蒸馏：用与 Full-PR 完全相同的特征族去拟合 NN 自身的预测
    #     （训练集上拟合，测试集上量残差）。
    with torch.no_grad():
        pred_nn_tr = model(X_tr).numpy()
    F_tr_d = custom_features_full(X_tr.numpy(), max_order=max_order_full_pr,
                                  include_special=include_special)
    F_te_d = custom_features_full(X_te_np, max_order=max_order_full_pr,
                                  include_special=include_special)
    distill = LinearRegression().fit(F_tr_d, pred_nn_tr)
    res["distill_resid"] = mean_squared_error(pred_nn, distill.predict(F_te_d))

    # 5. 模型对比输出（论文指标，表格化）
    if verbose:
        print(f"\n  ┌{'─' * 66}")
        print(f"  │ 模型对比（缩放空间内）")
        print(f"  ├{'─' * 66}")
        print(f"  │ {'模型':<12}{'MSE vs y':>12}{'MSE vs NN':>14}{'R² vs y':>10}")
        print(f"  │ {'NN':<12}{nn_mse:>12.6f}{'—':>14}"
              f"{r2_score(y_te_np, pred_nn):>10.4f}")
        if pred_tpr is not None:
            print(f"  │ {'Taylor-PR':<12}{res['tpr_mse_vs_y']:>12.6f}"
                  f"{res['tpr_mse_vs_nn']:>14.3e}"
                  f"{r2_score(y_te_np, pred_tpr):>10.4f}   ← MSE vs NN 为论文核心指标")
        print(f"  │ {'Full-PR':<12}{res['fpr_mse_vs_y']:>12.6f}"
              f"{res['fpr_mse_vs_nn']:>14.3e}"
              f"{r2_score(y_te_np, pred_fpr):>10.4f}   (max_order={max_order_full_pr})")
        print(f"  ├{'─' * 66}")
        print(f"  │ （参考，非对比标准）蒸馏残差 NN→PR族: "
              f"{res['distill_resid']:.3e}")
        print(f"  └{'─' * 66}")

    # 6. 模型两两之间的几何距离（按原项目方法）
    preds = {"NN": pred_nn}
    if pred_tpr is not None:
        preds["Taylor-PR"] = pred_tpr
    if pred_ltpr is not None:
        preds["LayerTaylor-PR"] = pred_ltpr
    preds["Full-PR"] = pred_fpr

    names = list(preds.keys())
    if verbose:
        print(f"\n  ┌{'─' * 66}")
        print(f"  │ 模型间几何距离（预测向量 Mahalanobis / 响应曲面 Kendall shape）")
        print(f"  ├{'─' * 66}")
        print(f"  │ {'模型对':<26}{'Mahalanobis':>13}{'Procrustes':>15}")
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            a, b = names[i], names[j]
            md = mahalanobis(preds[a], preds[b],
                             shrinkage='ledoitwolf', ridge=1e-6)
            sd = surface_shape_distance(X_te_np, preds[a], preds[b])
            res[f"mahal[{a}|{b}]"] = md
            res[f"shape[{a}|{b}]"] = sd
            if verbose:
                print(f"  │ {a + ' ↔ ' + b:<26}{_fmt(md, 13)}{_fmt(sd, 15)}")
    if verbose:
        print(f"  └{'─' * 66}")

    # 6b. 外推区的同等地位对比（论文 3.3 节"extended range"的数值版）：
    #     在数据缩放区间的 2 倍范围内均匀采样，比较两个模型在数据之外
    #     学到的函数是否一致。同分布上两个好模型必然接近（三角不等式），
    #     真正区分"学到同一个函数"与"恰好都拟合了数据"的是外推行为。
    lo, hi = scaling
    rng_ext = np.random.default_rng(data_seed + 10_000)
    X_ext = rng_ext.uniform(2 * lo, 2 * hi, size=(200, p))
    with torch.no_grad():
        pred_nn_ext = model(torch.tensor(X_ext, dtype=torch.float32)).numpy()
    pred_fpr_ext = fpr_model.predict(
        custom_features_full(X_ext, max_order=max_order_full_pr,
                             include_special=include_special))
    res["nnfpr_mse_ext"]   = mean_squared_error(pred_nn_ext, pred_fpr_ext)
    res["nnfpr_shape_ext"] = surface_shape_distance(X_ext, pred_nn_ext,
                                                    pred_fpr_ext)
    if pred_tpr is not None:
        pred_tpr_ext = eval_taylor_polynomial(
            *nn_to_taylor_polynomial(model, activation, q), X_ext)
        res["nntpr_shape_ext"] = surface_shape_distance(X_ext, pred_nn_ext,
                                                        pred_tpr_ext)
    if verbose:
        print(f"\n  ┌{'─' * 66}")
        print(f"  │ 外推区对比（均匀采样于 [{2*lo:.0f}, {2*hi:.0f}]^{p}，数据区间的 2 倍）")
        print(f"  ├{'─' * 66}")
        print(f"  │ NN ↔ Full-PR : MSE={res['nnfpr_mse_ext']:.3e}  "
              f"shape={res['nnfpr_shape_ext']:.4f}"
              f"   （数据区内 shape={res.get('shape[NN|Full-PR]', float('nan')):.4f}）")
        if 'nntpr_shape_ext' in res:
            print(f"  │ NN ↔ Taylor-PR: shape={res['nntpr_shape_ext']:.4f}"
                  f"   （数据区内 shape={res.get('shape[NN|Taylor-PR]', float('nan')):.4f}）")
        print(f"  └{'─' * 66}")

    # 7. NN 内部：相邻层 shape 距离 + 输入↔输出 shape 距离（新指标）
    _, shape_means, act_shape_means, io_shapes, lin_cc, act_cc = epoch_layer_distances(
        X_input=X_tr.numpy(), verbose=verbose)
    if shape_means:
        last_ep = max(shape_means.keys())
        res["adj_shape_mean_last"] = shape_means[last_ep]
        res["adj_shape_mean_avg"]  = float(np.nanmean(list(shape_means.values())))
        res["act_shape_mean_last"] = act_shape_means[last_ep]
        res["act_shape_mean_avg"]  = float(np.nanmean(list(act_shape_means.values())))
        res["io_shape_last"]       = io_shapes[last_ep]
        res["io_shape_avg"]        = float(np.nanmean(list(io_shapes.values())))

    # 7b. 训练轨迹：MSE 与形状距离并排（方便对比性能与几何演化）
    if verbose and history and shape_means:
        print(f"\n  ┌{'─' * 72}")
        print(f"  │ 训练轨迹：MSE 与形状距离对照")
        print(f"  ├{'─' * 72}")
        print(f"  │ {'Epoch':>6}{'Train MSE':>13}{'Test MSE':>13}"
              f"{'邻层均值':>11}{'激活对(非线性)':>15}{'输入↔输出':>12}")
        for ep in sorted(shape_means.keys()):
            rec = history.get(ep, {})
            tr = rec.get("train_mse", float("nan"))
            te = rec.get("test_mse",  float("nan"))
            print(f"  │ {ep:>6}{_fmt(tr, 13)}{_fmt(te, 13)}"
                  f"{_fmt(shape_means[ep], 11, 4)}"
                  f"{_fmt(act_shape_means[ep], 15, 4)}{_fmt(io_shapes[ep], 12, 4)}")
        print(f"  └{'─' * 72}")

    # 7c. 逐样本成对马氏距离（原项目流程）
    if lin_cc:
        last_ep = max(lin_cc.keys())
        _, stats = paired_md_from_first(
            lin_cc[last_ep], act_cc.get(last_ep, {}),
            source_name="model.0", align="auto",
            shrinkage="ledoitwolf", ridge=1e-6,
        )
        if verbose:
            print(f"\n  ┌{'─' * 66}")
            print(f"  │ 逐样本成对马氏距离（Epoch {last_ep}，以 model.0 为参照）")
            print(f"  ├{'─' * 66}")
            print(f"  │ {'目标层':<14}{'mean':>10}{'std':>10}")
            for layer, (mu, sd) in stats.items():
                print(f"  │ {layer:<14}{mu:>10.4f}{sd:>10.4f}")
            print(f"  └{'─' * 66}")

    return res


# ─────────────────────────────────────────────
# 重复模拟：复现论文 3.2 节的 MSE 分布研究（缩小版）
# ─────────────────────────────────────────────
def repeat_study(case, activation, h1, q, scaling, repeats, epochs,
                 include_special=True):
    """
    重复模拟 + 相关性研究：数值检验
        「NN 与 PR 的几何距离近 ⇒ 两者的拟合效果接近」

    度量约定：
      - 性能差距 gap = |RMSE_NN − RMSE_PR|（绝对差）。
        不用对数比值：当两者 MSE 都贴近噪声地板时比值发散，
        而几何距离约束的是加法结构上的绝对差（三角不等式）。
      - 距离一：RMSE(NN, PR) 预测间距离 —— gap ≤ 该距离是定理
        （三角不等式），程序对每次运行做自检；
      - 距离二：Procrustes 形状距离（模掉平移/缩放）—— 与 gap 的关系
        没有数学保证，是本研究真正要检验的经验命题；
      - 距离三：Mahalanobis（预测向量）。
    输出 Spearman 相关并导出 CSV（repeat_corr_<case>_<act>.csv）。
    """
    from scipy.stats import spearmanr

    print(f"\n{'#' * 78}")
    print(f"相关性研究：{case}, {activation}, h1={h1}, q={q}, repeats={repeats}")
    print(f"{'#' * 78}")
    print(f"{'seed':>4}{'RMSE(NN,PR)':>13}{'shape(内)':>11}{'mahal':>10}"
          f"{'gap|ΔRMSE|':>12}{'三角自检':>9}")

    rows = []
    for r in range(repeats):
        res = run_experiment(case=case, activation=activation,
                             hidden_layers=(h1,), q=q, scaling=scaling,
                             epochs=epochs, seed=r,
                             include_special=include_special, verbose=False)
        rmse_nn  = math.sqrt(res["nn_mse"])
        rmse_fpr = math.sqrt(res["fpr_mse_vs_y"])
        gap      = abs(rmse_nn - rmse_fpr)
        d_rmse   = math.sqrt(res["fpr_mse_vs_nn"])
        d_shape  = res.get("shape[NN|Full-PR]", float("nan"))
        d_mahal  = res.get("mahal[NN|Full-PR]", float("nan"))
        ok = "✓" if gap <= d_rmse + 1e-9 else "✗ 违反!"
        rows.append(dict(seed=r, d_rmse=d_rmse, d_shape=d_shape,
                         d_mahal=d_mahal, gap=gap,
                         rmse_nn=rmse_nn, rmse_fpr=rmse_fpr,
                         d_shape_ext=res.get("nnfpr_shape_ext", float("nan")),
                         tpr_gap=(abs(rmse_nn - math.sqrt(res["tpr_mse_vs_y"]))
                                  if "tpr_mse_vs_y" in res else float("nan")),
                         tpr_d_shape=res.get("shape[NN|Taylor-PR]", float("nan"))))
        print(f"{r:>4}{d_rmse:>13.4e}{d_shape:>11.4f}{d_mahal:>10.4f}"
              f"{gap:>12.4e}{ok:>9}")

    import csv
    fname = f"repeat_corr_{case}_{activation}_h{h1}.csv"
    with open(fname, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)

    def _corr(xs, ys, label):
        pairs = [(x, y) for x, y in zip(xs, ys)
                 if not (math.isnan(x) or math.isnan(y))]
        if len(pairs) < 3:
            print(f"  {label:<34}: 有效样本不足"); return
        rho, pval = spearmanr([p[0] for p in pairs], [p[1] for p in pairs])
        print(f"  {label:<34}: Spearman ρ = {rho:>7.3f}   p = {pval:.4f}"
              f"   (n={len(pairs)})")

    gaps = [r["gap"] for r in rows]
    print(f"\n  ── 距离 vs 性能差距 |ΔRMSE| 的 Spearman 相关 ──")
    _corr([r["d_rmse"]  for r in rows], gaps, "RMSE(NN,PR)（定理保证上界）")
    _corr([r["d_shape"] for r in rows], gaps, "Procrustes 内（经验命题★）")
    _corr([r["d_mahal"] for r in rows], gaps, "Mahalanobis")
    _corr([r["d_shape_ext"] for r in rows], gaps, "Procrustes 外推区")
    tg = [r["tpr_gap"] for r in rows]
    _corr([r["tpr_d_shape"] for r in rows], tg, "NN↔Taylor-PR shape vs 其gap")

    viol = sum(1 for r in rows if r["gap"] > r["d_rmse"] + 1e-9)
    print(f"\n  三角不等式自检：{repeats - viol}/{repeats} 通过"
          f"（gap ≤ RMSE(NN,PR) 应恒成立）")
    print(f"  明细已导出：{fname}（可用于画散点图）")
    return rows


# ─────────────────────────────────────────────
# 大规模蒙特卡洛模拟 + Excel 导出
# ─────────────────────────────────────────────
def _flatten_result(r):
    """把一次 run_experiment 的 res 摊平成单行 dict（只保留标量指标）。"""
    row = {
        "case": r["case"], "activation": r["activation"],
        "hidden": ",".join(str(x) for x in r["h"]),
        "n_hidden": len(r["h"]), "q": r["q"], "seed": r["seed"],
        "special": bool(r.get("fpr_special", True)),
        "NN_mse_vs_y":  r.get("nn_mse"),
        "FPR_mse_vs_y": r.get("fpr_mse_vs_y"),
        "FPR_mse_vs_NN": r.get("fpr_mse_vs_nn"),
        "TPR_mse_vs_NN": r.get("tpr_mse_vs_nn", float("nan")),
        "TPR_mse_vs_y":  r.get("tpr_mse_vs_y", float("nan")),
        "LTPR_mse_vs_NN": r.get("ltpr_mse_vs_nn", float("nan")),
        "LTPR_mse_vs_y":  r.get("ltpr_mse_vs_y", float("nan")),
        "LTPR_n_terms":   r.get("ltpr_n_terms", float("nan")),
        "shape_NN_LTPR":  r.get("shape[NN|LayerTaylor-PR]", float("nan")),
        "shape_LTPR_FPR": r.get("shape[LayerTaylor-PR|Full-PR]", float("nan")),
        "u_coverage":    r.get("u_coverage", float("nan")),
        # 几何指标
        "act_pair_mean":      r.get("act_shape_mean_last", float("nan")),
        "act_io_scaled_mean": r.get("act_io_scaled_mean", float("nan")),
        "act_io_unscaled_mean": r.get("act_io_unscaled_mean", float("nan")),
        "act_io_scaled_sum":  r.get("act_io_scaled_sum", float("nan")),
        "act_io_unscaled_sum": r.get("act_io_unscaled_sum", float("nan")),
        "io_shape":           r.get("io_shape_last", float("nan")),
        "shape_NN_FPR_in":    r.get("shape[NN|Full-PR]", float("nan")),
        "shape_NN_FPR_ext":   r.get("nnfpr_shape_ext", float("nan")),
        "mahal_NN_FPR":       r.get("mahal[NN|Full-PR]", float("nan")),
        "distill_resid":      r.get("distill_resid", float("nan")),
    }
    # 派生：性能差 |RMSE_NN − RMSE_FPR| 与 log10 比值
    if r.get("nn_mse", 0) > 0 and r.get("fpr_mse_vs_y", 0) > 0:
        row["abs_rmse_gap"] = abs(math.sqrt(r["nn_mse"])
                                  - math.sqrt(r["fpr_mse_vs_y"]))
        row["NN_FPR_log10"] = math.log10(r["nn_mse"] / r["fpr_mse_vs_y"])
    else:
        row["abs_rmse_gap"] = float("nan")
        row["NN_FPR_log10"] = float("nan")
    nn = row.get("NN_mse_vs_y")
    if nn is not None and nn > 0:
        for short, mse_col in [
            ("FPR", "FPR_mse_vs_y"),
            ("TPR", "TPR_mse_vs_y"),
            ("LTPR", "LTPR_mse_vs_y"),
        ]:
            mv = row.get(mse_col)
            if mv is not None and not math.isnan(mv):
                row[f"{short}_mse_delta_vs_NN"] = mv - nn
                row[f"{short}_improve_vs_NN_pct"] = 100.0 * (nn - mv) / nn
                row[f"{short}_better_than_NN"] = mv < nn
            else:
                row[f"{short}_mse_delta_vs_NN"] = float("nan")
                row[f"{short}_improve_vs_NN_pct"] = float("nan")
                row[f"{short}_better_than_NN"] = float("nan")
    return row


def monte_carlo(epochs=800, seeds=30, out_path="monte_carlo_results.xlsx",
                cases=None, activations=("softplus", "tanh", "sigmoid"),
                depths=((4,), (8, 4), (16, 8, 4)), specials=(True, False),
                q=3):
    """
    大规模蒙特卡洛：对每个 (case × activation × depth × special) 配置跑 `seeds`
    个不同随机种子，收集所有指标，导出 Excel：
      - Sheet 'raw'      ：每次模拟一行（全部指标）
      - Sheet 'summary'  ：按配置聚合的 均值 / 标准差
      - Sheet 'correlations'：跨全部样本，几何距离 vs 性能/接近度 的相关
    """
    import pandas as pd
    from scipy.stats import spearmanr, pearsonr

    if cases is None:
        cases = ["paper_poly4", "smooth_nonlinear_rand"]

    configs = [(c, a, d, sp) for c in cases for a in activations
               for d in depths for sp in specials]
    total = len(configs) * seeds
    print(f"蒙特卡洛：{len(configs)} 配置 × {seeds} seeds = {total} 次模拟")

    rows, done = [], 0
    for (case, act, depth, sp) in configs:
        for s in range(seeds):
            r = run_experiment(case=case, activation=act, hidden_layers=depth,
                               q=q, epochs=epochs, seed=s,
                               include_special=sp, verbose=False)
            rows.append(_flatten_result(r))
            done += 1
            if done % 10 == 0 or done == total:
                print(f"  进度 {done}/{total}")

    df = pd.DataFrame(rows)

    # 聚合：按配置算 均值/标准差
    group_keys = ["case", "activation", "hidden", "special"]
    metric_cols = [c for c in df.columns if c not in
                   group_keys + ["n_hidden", "q", "seed"]]
    summary = df.groupby(group_keys)[metric_cols].agg(["mean", "std"])
    summary.columns = [f"{m}_{stat}" for m, stat in summary.columns]
    summary = summary.reset_index()

    # 相关性：几何距离 vs (接近度 shape_NN_FPR_in / 性能差 abs_rmse_gap)
    def safe_corr(x, y):
        m = df[[x, y]].dropna()
        if len(m) < 5:
            return (float("nan"), float("nan"), float("nan"), len(m))
        rho, p_s = spearmanr(m[x], m[y])
        r_p, p_p = pearsonr(m[x], m[y])
        return (rho, p_s, r_p, len(m))

    corr_pairs = [
        ("act_io_scaled_mean",   "shape_NN_FPR_in"),
        ("act_io_unscaled_mean", "shape_NN_FPR_in"),
        ("act_pair_mean",        "shape_NN_FPR_in"),
        ("act_io_scaled_mean",   "shape_NN_FPR_ext"),
        ("act_io_scaled_mean",   "abs_rmse_gap"),
        ("act_io_scaled_mean",   "TPR_mse_vs_NN"),
        ("io_shape",             "shape_NN_FPR_in"),
        ("shape_NN_FPR_in",      "abs_rmse_gap"),
    ]
    corr_rows = []
    for x, y in corr_pairs:
        rho, ps, rp, n = safe_corr(x, y)
        corr_rows.append({"x": x, "y": y, "spearman_rho": rho,
                          "spearman_p": ps, "pearson_r": rp, "n": n})
    corr_df = pd.DataFrame(corr_rows)

    # 写 Excel（三个 sheet）
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="raw", index=False)
        summary.to_excel(xl, sheet_name="summary", index=False)
        corr_df.to_excel(xl, sheet_name="correlations", index=False)

    # 简单加粗表头 + 冻结首行 + 列宽
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial")
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None),
                        default=8)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 24)
    wb.save(out_path)

    print(f"\n已导出：{out_path}")
    print(f"  raw          : {len(df)} 行 × {len(df.columns)} 列")
    print(f"  summary      : {len(summary)} 个配置的均值/标准差")
    print(f"  correlations : {len(corr_df)} 组相关性检验")
    print(f"\n关键相关（跨全部 {len(df)} 样本）：")
    for _, cr in corr_df.iterrows():
        print(f"  {cr['x']:<22} vs {cr['y']:<18}: "
              f"ρ={cr['spearman_rho']:>6.3f} (p={cr['spearman_p']:.3g}, n={int(cr['n'])})")
    return df, summary, corr_df


# ─────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────
def _grid_worker(case, act, hidden, special, q, epochs, data_seed, init_seed):
    """单个 (配置, init_seed) 的工作单元——供多进程并行调用。
    必须是模块顶层函数，否则无法被 pickle 传到子进程。
    每个子进程有独立的全局 hook 状态(reset_capture 在 run_experiment 内调用)，
    因此进程间互不干扰。"""
    r = run_experiment(case=case, activation=act, hidden_layers=hidden,
                       q=q, epochs=epochs, data_seed=data_seed,
                       init_seed=init_seed, include_special=special,
                       verbose=False)
    row = _flatten_result(r)
    row["well_specified"] = (
        (case == "paper_poly2" and len(hidden) == 2) or
        (case == "paper_poly3" and len(hidden) == 3) or
        (case == "paper_poly4" and len(hidden) == 4))
    return row


def default_grid(epochs, repeats=50, data_seed=0, out_prefix="results",
                 n_jobs=1):
    """
    重复模拟实验组（论文式：固定数据、扫网络初始化）。
      - 每个配置用同一份数据（data_seed 固定），扫 `repeats` 个不同的
        网络初始化 init_seed，反映"同数据、不同局部极小"的训练随机性。
      - 数据集：
          paper_poly4 (4阶)     ：欠设定（>最大隐层数3），永不命中
          paper_poly3 (3阶)     ：3 隐层时【正好命中】（well-specified 对照）
          smooth_nonlinear_rand ：非多项式，严格欠设定
          smooth_nonlinear      ：固定频率非多项式对照
      - 深度 (4,)/(8,4)/(16,8,4)，Full-PR 含/不含特殊项各一组。
    输出：聚合汇总表、相关性摘要表（你的提议C）、箱线图+散点图、CSV。
    """
    import pandas as pd
    from scipy.stats import spearmanr

    depth_variants = [(4,), (8, 4), (16, 8, 4)]
    base_configs = (
        [("paper_poly4", a) for a in ("softplus", "tanh", "sigmoid")]
        + [("paper_poly3", a) for a in ("softplus", "tanh", "sigmoid")]
        + [("smooth_nonlinear_rand", a) for a in ("softplus", "tanh", "sigmoid")]
        + [("smooth_nonlinear", "softplus")]
    )

    def is_well_specified(case, n_hidden):
        return (case == "paper_poly2" and n_hidden == 2) or \
               (case == "paper_poly3" and n_hidden == 3) or \
               (case == "paper_poly4" and n_hidden == 4)

    plan = [(c, a, h, sp) for (c, a) in base_configs
            for h in depth_variants for sp in (True, False)]
    total = len(plan) * repeats
    print(f"重复模拟：{len(base_configs)} 基础配置 × {len(depth_variants)} 深度 "
          f"× 2 特殊项 × {repeats} 初始化 seed = {total} 次")
    print(f"并行进程数 n_jobs={n_jobs}"
          f"{'（串行）' if n_jobs == 1 else ''}")

    tasks = [(case, act, hidden, special, 3, epochs, data_seed, init_seed)
             for (case, act, hidden, special) in plan
             for init_seed in range(repeats)]

    if n_jobs == 1:
        rows = []
        for i, t in enumerate(tasks):
            rows.append(_grid_worker(*t))
            if (i + 1) % 25 == 0 or (i + 1) == total:
                print(f"  进度 {i + 1}/{total}")
    else:
        from joblib import Parallel, delayed
        rows = Parallel(n_jobs=n_jobs, verbose=10)(
            delayed(_grid_worker)(*t) for t in tasks)

    df = pd.DataFrame(rows)

    # ── 1) 聚合汇总表（均值±标准差）──
    gk = ["case", "activation", "hidden", "special", "well_specified"]
    agg_cols = ["NN_mse_vs_y", "FPR_mse_vs_y", "FPR_mse_vs_NN", "TPR_mse_vs_NN",
                "LTPR_mse_vs_NN", "shape_NN_LTPR",
                "act_io_scaled_mean", "act_io_unscaled_mean",
                "shape_NN_FPR_in", "shape_NN_FPR_ext", "NN_FPR_log10",
                "FPR_improve_vs_NN_pct", "TPR_improve_vs_NN_pct",
                "LTPR_improve_vs_NN_pct"]
    g = df.groupby(gk)[agg_cols].agg(["mean", "std"])

    print(f"\n{'═' * 178}")
    print(f"聚合汇总（每格 {repeats} 次重复 均值±标准差；★=well-specified 正好命中；"
          f"LTPR=逐层泰勒展开多项式）")
    print(f"{'═' * 178}")
    print(f"{'case':<22}{'act':<9}{'hid':<8}{'特殊':<5}{'命中':<5}"
          f"{'NN_mse':>16}{'FPR_mse':>16}{'FPR vs NN':>16}"
          f"{'LTPR vs NN':>16}{'形NN-LTPR':>16}"
          f"{'u→gu缩放':>14}{'形NN-FPR(内)':>16}{'形(外)':>16}")
    print('─' * 178)
    prev_case = None
    for idx, sub in g.iterrows():
        case, act, hid, sp, ws = idx
        if prev_case is not None and case != prev_case:
            print('─' * 178)
        prev_case = case
        def ms(col):
            m, s = sub[(col, "mean")], sub[(col, "std")]
            if math.isnan(m):
                return f"{'—':>14}"
            return f"{m:.4f}±{s:.4f}"
        print(f"{case:<22}{act:<9}{hid:<8}{'✓' if sp else '✗':<5}"
              f"{'★' if ws else ' ':<5}"
              f"{ms('NN_mse_vs_y'):>16}{ms('FPR_mse_vs_y'):>16}"
              f"{ms('FPR_mse_vs_NN'):>16}"
              f"{ms('LTPR_mse_vs_NN'):>16}{ms('shape_NN_LTPR'):>16}"
              f"{ms('act_io_scaled_mean'):>14}"
              f"{ms('shape_NN_FPR_in'):>16}{ms('shape_NN_FPR_ext'):>16}")
    print('═' * 178)

    # ── 2) 相关性摘要表（你的提议C）──
    corr_pairs = [
        ("act_io_scaled_sum",   "shape_NN_FPR_in",  "激活改变总值(缩放) → NN-PR接近度(内)"),
        ("act_io_unscaled_sum", "shape_NN_FPR_in",  "激活改变总值(未缩) → NN-PR接近度(内)"),
        ("act_io_scaled_sum",   "shape_NN_FPR_ext", "激活改变总值(缩放) → NN-PR接近度(外推)"),
        ("act_io_scaled_sum",   "abs_rmse_gap",     "激活改变总值(缩放) → 性能差距"),
        ("act_pair_mean",       "shape_NN_FPR_in",  "相邻层非线性 → NN-PR接近度(内)"),
        ("shape_NN_FPR_in",     "abs_rmse_gap",     "NN-PR(Full)接近度 → 性能差距"),
        # LayerTaylor-PR 相关：逐层泰勒多项式与 NN 的接近度
        ("shape_NN_LTPR",       "abs_rmse_gap",     "NN-LayerTaylorPR接近度 → 性能差距"),
        ("act_io_scaled_sum",   "shape_NN_LTPR",    "激活改变总值(缩放) → NN-LayerTaylorPR接近度"),
        ("shape_NN_LTPR",       "shape_NN_FPR_in",  "NN-LayerTaylorPR ↔ NN-FullPR 接近度"),
        # 逐层泰勒展开多项式 与 Full-PR 的直接关系（你要的图）
        ("shape_LTPR_FPR",      "abs_rmse_gap",     "LayerTaylorPR-FullPR接近度 → 性能差距"),
        ("shape_LTPR_FPR",      "shape_NN_FPR_in",  "LayerTaylorPR-FullPR ↔ NN-FullPR 接近度"),
    ]
    corr_records = []
    groups = [("ALL", df)] + [(c, df[df["case"] == c]) for c in df["case"].unique()]
    for gname, gdf in groups:
        for x, y, label in corr_pairs:
            m = gdf[[x, y]].dropna()
            if len(m) >= 5 and m[x].std() > 0 and m[y].std() > 0:
                rho, pval = spearmanr(m[x], m[y])
            else:
                rho, pval = float("nan"), float("nan")
            corr_records.append({"group": gname, "relation": label, "x": x, "y": y,
                                 "spearman_rho": rho, "p_value": pval, "n": len(m)})
    corr_df = pd.DataFrame(corr_records)

    print(f"\n{'═' * 96}")
    print("相关性摘要（Spearman ρ；几何距离 vs 接近度/性能）")
    print(f"{'═' * 96}")
    print(f"{'数据集':<24}{'关系':<40}{'ρ':>8}{'p':>10}{'n':>6}")
    print('─' * 96)
    prev_g = None
    for _, cr in corr_df.iterrows():
        if prev_g is not None and cr["group"] != prev_g:
            print('─' * 96)
        prev_g = cr["group"]
        sig = "*" if (not math.isnan(cr["p_value"]) and cr["p_value"] < 0.05) else " "
        rho_s = f"{cr['spearman_rho']:.3f}" if not math.isnan(cr['spearman_rho']) else "—"
        p_s = f"{cr['p_value']:.3g}" if not math.isnan(cr['p_value']) else "—"
        print(f"{cr['group']:<24}{cr['relation']:<40}{rho_s:>7}{sig}{p_s:>10}{cr['n']:>6}")
    print(f"{'═' * 96}")
    print("(* 表示 p<0.05)")

    # ── 3) 可视化 ──
    try:
        _make_plots(df, corr_pairs, out_prefix)
        plotted = True
    except Exception as e:
        print(f"绘图跳过（{e}）")
        plotted = False

    # ── 4) CSV 导出 ──
    raw_csv = f"{out_prefix}_raw.csv"
    corr_csv = f"{out_prefix}_correlations.csv"
    df.to_csv(raw_csv, index=False, encoding="utf-8-sig")
    corr_df.to_csv(corr_csv, index=False, encoding="utf-8-sig")
    print(f"\n已导出：")
    print(f"  {raw_csv}          （{len(df)} 行原始记录）")
    print(f"  {corr_csv} （相关性摘要）")
    if plotted:
        cases = sorted(df["case"].unique())
        print(f"  按数据集分别出图（每个数据集 3 张：_geometry.png / _mse.png / _nn_baseline.png）：")
        for c in cases:
            print(f"    {out_prefix}_{c}_geometry.png, {out_prefix}_{c}_mse.png, "
                  f"{out_prefix}_{c}_nn_baseline.png")
        print(f"  {out_prefix}_overview_boxplot.png （总览箱线图）")
    return df, corr_df


def _make_plots(df, corr_pairs, out_prefix):
    """
    按数据集分别出图，避免不同数据集混在一张图里互相掩盖（Simpson 悖论）。
    每个数据集生成三张图：
      <prefix>_<case>_geometry.png  —— 几何距离 vs 接近度/性能（散点，按激活着色）
      <prefix>_<case>_mse.png       —— 各测量指标 vs MSE、以及模型间表现关系
      <prefix>_<case>_nn_baseline.png —— 以 NN 为基准，显示哪个模型更好
    另外保留一张总览箱线图。
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    acts = ["softplus", "tanh", "sigmoid"]
    colors = {"softplus": "#2a9d8f", "tanh": "#e76f51", "sigmoid": "#264653"}
    compare_models = [
        ("FPR", "FullPR"),
        ("TPR", "Taylor-PR"),
        ("LTPR", "LayerTaylor-PR"),
    ]

    en_label = {
        ("act_io_scaled_sum", "shape_NN_FPR_in"):  "act-change(sum,scaled) -> NN-FullPR closeness(in)",
        ("act_io_unscaled_sum", "shape_NN_FPR_in"): "act-change(sum,unscaled) -> NN-FullPR closeness(in)",
        ("act_io_scaled_sum", "shape_NN_FPR_ext"):  "act-change(sum,scaled) -> closeness(extrap)",
        ("act_io_scaled_sum", "abs_rmse_gap"):      "act-change(sum,scaled) -> perf gap",
        ("act_pair_mean", "shape_NN_FPR_in"):       "adj-layer nonlin -> NN-FullPR closeness(in)",
        ("shape_NN_FPR_in", "abs_rmse_gap"):        "NN-FullPR closeness -> perf gap",
        ("shape_NN_LTPR", "abs_rmse_gap"):          "NN-LayerTaylorPR closeness -> perf gap",
        ("act_io_scaled_sum", "shape_NN_LTPR"):     "act-change -> NN-LayerTaylorPR closeness",
        ("shape_NN_LTPR", "shape_NN_FPR_in"):       "NN-LayerTaylorPR vs NN-FullPR closeness",
        ("shape_LTPR_FPR", "abs_rmse_gap"):         "LayerTaylorPR-FullPR closeness -> perf gap",
        ("shape_LTPR_FPR", "shape_NN_FPR_in"):      "LayerTaylorPR-FullPR vs NN-FullPR",
    }

    def scatter_grid(sub_df, pairs, title, path):
        n = len(pairs)
        ncol = 3
        nrow = (n + ncol - 1) // ncol
        fig, axes = plt.subplots(nrow, ncol, figsize=(15, 4.3 * nrow))
        axes_flat = axes.ravel() if hasattr(axes, "ravel") else [axes]
        for ax, (x, y, _lab) in zip(axes_flat, pairs):
            if x not in sub_df.columns or y not in sub_df.columns:
                ax.set_visible(False); continue
            for a in acts:
                s = sub_df[sub_df["activation"] == a][[x, y]].dropna()
                ax.scatter(s[x], s[y], s=16, alpha=0.5, color=colors[a], label=a)
            ax.set_xlabel(x); ax.set_ylabel(y)
            ax.set_title(en_label.get((x, y), f"{x} vs {y}"), fontsize=9)
            ax.grid(alpha=0.3); ax.legend(fontsize=7)
        for ax in axes_flat[n:]:
            ax.set_visible(False)
        fig.suptitle(title, fontsize=13)
        fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)

    # 几何距离 / 接近度 / 性能 的散点对
    geom_pairs = corr_pairs

    # MSE & 模型间表现关系 的散点对（直接调查各指标与 MSE 的关系，以及模型互比）
    mse_pairs = [
        ("act_io_scaled_sum", "NN_mse_vs_y",   "act-change -> NN MSE(vs y)"),
        ("shape_NN_FPR_in",   "FPR_mse_vs_NN", "NN-FullPR closeness -> FullPR MSE(vs NN)"),
        ("shape_NN_LTPR",     "LTPR_mse_vs_NN","NN-LayerTaylorPR closeness -> LTPR MSE(vs NN)"),
        ("NN_mse_vs_y",       "FPR_mse_vs_y",  "NN MSE vs FullPR MSE (both vs y)"),
        ("FPR_mse_vs_NN",     "LTPR_mse_vs_NN","FullPR vs LayerTaylorPR (both MSE vs NN)"),
        ("NN_mse_vs_y",       "abs_rmse_gap",  "NN MSE -> NN-FullPR perf gap"),
    ]
    mse_label = {p[:2]: p[2] for p in mse_pairs}

    def mse_grid(sub_df, title, path):
        fig, axes = plt.subplots(2, 3, figsize=(15, 9))
        for ax, (x, y, _l) in zip(axes.ravel(), mse_pairs):
            if x not in sub_df.columns or y not in sub_df.columns:
                ax.set_visible(False); continue
            for a in acts:
                s = sub_df[sub_df["activation"] == a][[x, y]].dropna()
                ax.scatter(s[x], s[y], s=16, alpha=0.5, color=colors[a], label=a)
            # 对“两模型 MSE 互比”那张加 y=x 参考线
            if (x, y) == ("NN_mse_vs_y", "FPR_mse_vs_y"):
                lim = max(sub_df[x].max(), sub_df[y].max())
                ax.plot([0, lim], [0, lim], "k--", lw=0.8, alpha=0.6)
            ax.set_xlabel(x); ax.set_ylabel(y)
            ax.set_title(mse_label.get((x, y), f"{x} vs {y}"), fontsize=9)
            ax.grid(alpha=0.3); ax.legend(fontsize=7)
        fig.suptitle(title, fontsize=13)
        fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)

    def nn_baseline_grid(sub_df, title, path):
        rows = []
        for short, name in compare_models:
            pct_col = f"{short}_improve_vs_NN_pct"
            win_col = f"{short}_better_than_NN"
            if pct_col not in sub_df.columns or win_col not in sub_df.columns:
                continue
            for act in acts:
                s = sub_df[sub_df["activation"] == act][[pct_col, win_col]].dropna()
                if len(s) == 0:
                    continue
                rows.append((name, act, float(s[win_col].mean() * 100.0),
                             float(s[pct_col].median())))
        if not rows:
            return
        models = list(dict.fromkeys(r[0] for r in rows))
        x = np.arange(len(models))
        width = 0.24
        offsets = np.linspace(-width, width, len(acts))
        fig, axes = plt.subplots(1, 2, figsize=(13, 5), sharex=True)
        for i, act in enumerate(acts):
            win_vals, imp_vals = [], []
            for model in models:
                hit = [r for r in rows if r[0] == model and r[1] == act]
                win_vals.append(hit[0][2] if hit else np.nan)
                imp_vals.append(hit[0][3] if hit else np.nan)
            axes[0].bar(x + offsets[i], win_vals, width=width, color=colors[act],
                        alpha=0.75, label=act)
            axes[1].bar(x + offsets[i], imp_vals, width=width, color=colors[act],
                        alpha=0.75, label=act)
        axes[0].axhline(50, color="k", lw=0.8, ls="--", alpha=0.5)
        axes[0].set_ylabel("win rate vs NN (%)")
        axes[0].set_title("How often model MSE < NN MSE", fontsize=10)
        axes[1].axhline(0, color="k", lw=0.8, ls="--", alpha=0.6)
        axes[1].set_ylabel("median improvement vs NN (%)")
        axes[1].set_title("Positive means better than NN", fontsize=10)
        for ax in axes:
            ax.set_xticks(x)
            ax.set_xticklabels(models, rotation=15, ha="right")
            ax.grid(axis="y", alpha=0.3)
            ax.legend(fontsize=8)
        fig.suptitle(title, fontsize=13)
        fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)

    # ── 按数据集分别出图 ──
    for case in sorted(df["case"].unique()):
        sub = df[df["case"] == case]
        scatter_grid(sub, geom_pairs,
                     f"[{case}] Geometric distance vs closeness/performance",
                     f"{out_prefix}_{case}_geometry.png")
        mse_grid(sub,
                 f"[{case}] Measurements vs MSE & model-vs-model performance",
                 f"{out_prefix}_{case}_mse.png")
        nn_baseline_grid(sub,
                         f"[{case}] Models vs NN baseline",
                         f"{out_prefix}_{case}_nn_baseline.png")

    # ── 总览箱线图（按激活函数，跨全部数据集）──
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))
    for ax, col, title in [
        (axes[0], "act_io_scaled_sum", "activation change u->g(u) (sum)"),
        (axes[1], "shape_NN_FPR_in", "NN-FullPR shape dist (in)"),
        (axes[2], "abs_rmse_gap", "NN-FullPR perf gap")]:
        data = [df[df["activation"] == a][col].dropna().values for a in acts]
        bp = ax.boxplot(data, labels=acts, patch_artist=True, showfliers=True)
        for patch, a in zip(bp["boxes"], acts):
            patch.set_facecolor(colors[a]); patch.set_alpha(0.6)
        ax.set_title(title); ax.set_ylabel(col); ax.grid(alpha=0.3)
    fig.suptitle("Overview: metrics by activation (all datasets, all repeats)",
                 fontsize=13)
    fig.tight_layout(); fig.savefig(f"{out_prefix}_overview_boxplot.png", dpi=130)
    plt.close(fig)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Morala et al. 风格实验 + 几何距离测量")
    parser.add_argument("--case", type=str, default=None,
                        choices=["paper_poly2", "paper_poly3", "paper_poly4",
                                 "smooth_nonlinear", "smooth_nonlinear_rand"],
                        help="数据集；不指定则跑默认全部实验组")
    parser.add_argument("--activation", type=str, default="softplus",
                        choices=list(ACTIVATIONS.keys()),
                        help="激活函数接口")
    parser.add_argument("--h1", type=str, default="4",
                        help="隐层宽度，逗号分隔可指定多隐层，如 '4' 或 '8,4'。"
                             "Full-PR 的 max_order = 隐层层数；"
                             "Taylor-PR 仅在单隐层时构建")
    parser.add_argument("--q", type=int, default=3, help="泰勒展开截断阶数")
    parser.add_argument("--scaling", type=str, default="-1,1",
                        help="min-max 缩放区间，'-1,1' 或 '0,1'")
    parser.add_argument("--n", type=int, default=200, help="样本量（论文用 200）")
    parser.add_argument("--epochs", type=int, default=1500)
    parser.add_argument("--optimizer", type=str, default="rprop",
                        choices=["rprop", "adam"])
    parser.add_argument("--capture-every", type=int, default=150,
                        help="每隔多少 epoch 采集一次层输出用于几何距离")
    parser.add_argument("--no-special", action="store_true",
                        help="Full-PR 去掉 sin/e^x 特殊项，使用纯多项式（建议二）")
    parser.add_argument("--repeats", type=int, default=0,
                        help=">0 时做重复模拟 + 距离-性能相关性研究")
    parser.add_argument("--monte-carlo", type=int, default=0,
                        help=">0 时跑大规模蒙特卡洛（该数=每配置的 seed 数），导出 Excel")
    parser.add_argument("--excel-out", type=str, default="monte_carlo_results.xlsx",
                        help="蒙特卡洛 Excel 输出路径")
    parser.add_argument("--seed", type=int, default=0)
    parser.add_argument("--grid-repeats", type=int, default=50,
                        help="默认网格里每个配置的重复次数（扫初始化 seed）")
    parser.add_argument("--data-seed", type=int, default=0,
                        help="默认网格里固定的数据生成 seed")
    parser.add_argument("--out-prefix", type=str, default="results",
                        help="默认网格 CSV/图 的输出前缀")
    parser.add_argument("--n-jobs", type=int, default=1,
                        help="默认网格的并行进程数（-1=用全部 CPU 核心）")
    args = parser.parse_args()

    scaling = tuple(float(s) for s in args.scaling.split(","))
    hidden = tuple(int(s) for s in args.h1.split(","))
    include_special = not args.no_special

    if args.monte_carlo > 0:
        monte_carlo(epochs=args.epochs, seeds=args.monte_carlo,
                    out_path=args.excel_out)
    elif args.repeats > 0:
        repeat_study(case=args.case or "paper_poly4",
                     activation=args.activation, h1=hidden[0], q=args.q,
                     scaling=scaling, repeats=args.repeats, epochs=args.epochs,
                     include_special=include_special)
    elif args.case is None:
        default_grid(epochs=args.epochs, repeats=args.grid_repeats,
                     data_seed=args.data_seed, out_prefix=args.out_prefix,
                     n_jobs=args.n_jobs)
    else:
        run_experiment(case=args.case, activation=args.activation,
                       hidden_layers=hidden, q=args.q, scaling=scaling,
                       n=args.n, epochs=args.epochs, optimizer=args.optimizer,
                       capture_every=args.capture_every, seed=args.seed,
                       include_special=include_special)
